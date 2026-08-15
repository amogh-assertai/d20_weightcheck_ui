"""
app/main/history_routes.py
---------------------------
History table, Activity Details (with prev/next through the filtered set),
and saving the human review fields.
"""

import os
from datetime import datetime, timezone
from io import BytesIO

from bson import ObjectId
from bson.errors import InvalidId
from flask import (
    current_app, flash, redirect, render_template, request, send_file, url_for,
)
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.main import main_bp
from app.main.shared import base_context, query_activities
from app.utils.helpers import format_timestamp

ALLOWED_PER_PAGE = (25, 50, 100)
DEFAULT_PER_PAGE = 25

# Never export these:
#  - _id/image_path/raw_image_path: not useful in a spreadsheet, and the
#    image itself was explicitly out of scope
#  - mode: not wanted in the export
#  - api_id/api_image_path/uploaded_at: these are sync-bookkeeping fields
#    the LOCAL uploader script adds to its own copy after a successful
#    upload - they aren't meant to be business data, they just come through
#    as "extra fields, stored as-is" if a document happens to carry them.
EXPORT_EXCLUDED_FIELDS = {
    "_id", "image_path", "raw_image_path", "mode",
    "api_id", "api_image_path", "uploaded_at",
}

# Column headers that shouldn't just be the field name title-cased -
# mark_ocr_wrong is labelled "System error" everywhere else in the UI, so
# the export header should say the same thing, not "Mark Ocr Wrong".
EXPORT_HEADER_OVERRIDES = {
    "mark_ocr_wrong": "Mark System Error",
}

# Fields that hold a raw timestamp string and should be shown formatted
# (e.g. "24 Jul 2026, 22:28:48"), not MongoDB's raw ISO/BSON value.
EXPORT_DATETIME_FIELDS = {"timestamp", "created_at"}

# Preferred column order for the export - anything else found on a document
# (custom/future fields) is appended afterwards, sorted alphabetically, so
# nothing is ever silently dropped even if the schema grows.
EXPORT_PREFERRED_FIELD_ORDER = [
    "camera_id", "camera_name", "activity_number",
    "expected_order_number", "actual_order_number", "order_number_matching",
    "order_number_result", "order_number_reason",
    "expected_weight", "actual_weight", "weight_difference", "weight_difference_percent",
    "weight_result", "weight_reason",
    "validation_result", "validation_reason",
    "timestamp", "created_at",
    "mark_discuss", "mark_ocr_wrong", "mark_process_error", "review_comment", "result_reviewed",
]


@main_bp.route("/history")
def history():
    """
    History - filterable, searchable, sortable, paginated view of activity records.
    Filters (table / date range / result / ocr-wrong / has-comment) apply first,
    then search narrows the already-filtered set, then sort, then pagination.
    """
    db = current_app.extensions.get("mongo_db")
    error = None
    activities = []
    total_count = 0
    camera_options = []
    meta = {}

    try:
        per_page = int(request.args.get("per_page", DEFAULT_PER_PAGE))
    except ValueError:
        per_page = DEFAULT_PER_PAGE
    if per_page not in ALLOWED_PER_PAGE:
        per_page = DEFAULT_PER_PAGE

    try:
        page = max(int(request.args.get("page", 1)), 1)
    except ValueError:
        page = 1

    if db is None:
        error = "Database not configured."
    else:
        try:
            filtered, camera_options, meta = query_activities(db)
            total_count = len(filtered)

            start_idx = (page - 1) * per_page
            page_items = filtered[start_idx:start_idx + per_page]

            for doc in page_items:
                doc["_id"] = str(doc["_id"])
                doc["timestamp_display"] = format_timestamp(doc.get("timestamp"))

            activities = page_items
        except Exception as e:
            current_app.logger.error(f"Error fetching activities for history page: {e}")
            error = "Could not load activity history."

    total_pages = max((total_count + per_page - 1) // per_page, 1)

    context = base_context()
    context.update(meta)
    context.update({
        "activities": activities,
        "error": error,
        "camera_options": camera_options,
        "page": page,
        "per_page": per_page,
        "allowed_per_page": ALLOWED_PER_PAGE,
        "total_pages": total_pages,
        "total_count": total_count,
    })

    try:
        return render_template("main/history.html", **context)
    except Exception as e:
        current_app.logger.error(f"Error rendering history page: {e}")
        return "Something went wrong loading history.", 500


@main_bp.route("/history/export")
def export_history():
    """
    Export the CURRENTLY FILTERED set (same table/date/result/search/review
    filters as the History page, via the same query_activities() helper -
    not just the current page) to an .xlsx file. No image data - every
    other field on the activity, including the review fields (comments,
    marked-for-discussion, etc.), is included.

    The frontend shows a confirmation modal before hitting this route,
    warning that only the active filters' results will be exported - this
    route itself doesn't second-guess that; it exports whatever the query
    string says, in full (no pagination limit).
    """
    db = current_app.extensions.get("mongo_db")
    if db is None:
        return "Database not configured.", 503

    try:
        filtered, _camera_options, _meta = query_activities(db)
    except Exception as e:
        current_app.logger.error(f"Error building export data set: {e}")
        return "Something went wrong preparing the export.", 500

    # Column order: preferred fields first, then anything else found on any
    # document (alphabetical) - so a custom/future field is never silently
    # dropped, it just lands at the end instead of a fixed position.
    present_fields = set()
    for doc in filtered:
        present_fields.update(doc.keys())
    present_fields -= EXPORT_EXCLUDED_FIELDS

    if present_fields:
        ordered = [f for f in EXPORT_PREFERRED_FIELD_ORDER if f in present_fields]
        remaining = sorted(present_fields - set(ordered))
        columns = ordered + remaining
    else:
        # Zero matching activities - still produce a valid file with a
        # sensible header row (not a headerless/broken spreadsheet).
        columns = list(EXPORT_PREFERRED_FIELD_ORDER)

    wb = Workbook()
    ws = wb.active
    ws.title = "Activities"
    ws.append([
        EXPORT_HEADER_OVERRIDES.get(c, c.replace("_", " ").title())
        for c in columns
    ])

    for doc in filtered:
        row = []
        for col in columns:
            value = doc.get(col, "")
            if col in EXPORT_DATETIME_FIELDS and value:
                value = format_timestamp(value)
            elif isinstance(value, (dict, list)):
                value = str(value)  # keep exportable - Excel cells can't hold nested objects
            row.append(value)
        ws.append(row)

    for i, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, min(40, len(col) + 4))

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"activity_history_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@main_bp.route("/history/activity/<activity_id>")
def activity_detail(activity_id):
    """
    Activity details: prev/next navigation (within the same filtered+sorted set
    the person was browsing on History), evidence image in the middle, order/weight
    fields on the left, activity/reasoning fields on the right, and a review form
    below (mark for discussion, mark OCR wrong, free-text comment).
    """
    db = current_app.extensions.get("mongo_db")
    if db is None:
        return "Database not configured.", 503

    try:
        obj_id = ObjectId(activity_id)
    except InvalidId:
        return "Activity not found.", 404

    try:
        doc = db["all_activities"].find_one({"_id": obj_id})
    except Exception as e:
        current_app.logger.error(f"Error fetching activity '{activity_id}': {e}")
        return "Something went wrong loading this activity.", 500

    if doc is None:
        return "Activity not found.", 404

    # Compute prev/next within the same filtered+sorted set as History
    # (uses whatever filter query params are attached to this page's own URL).
    prev_id = next_id = None
    try:
        filtered, _, _ = query_activities(db)
        ids_in_order = [str(d["_id"]) for d in filtered]
        if activity_id in ids_in_order:
            idx = ids_in_order.index(activity_id)
            if idx > 0:
                prev_id = ids_in_order[idx - 1]
            if idx < len(ids_in_order) - 1:
                next_id = ids_in_order[idx + 1]
    except Exception as e:
        current_app.logger.warning(f"Could not compute prev/next for activity '{activity_id}': {e}")

    doc["_id"] = str(doc["_id"])
    doc["timestamp_display"] = format_timestamp(doc.get("timestamp"))

    # Build a servable URL for the evidence image, if it's inside UPLOAD_FOLDER
    image_url = None
    image_path = doc.get("image_path")
    if image_path:
        try:
            rel_path = os.path.relpath(image_path, current_app.config["UPLOAD_FOLDER"])
            if not rel_path.startswith(".."):
                image_url = url_for("main.serve_media", filepath=rel_path)
        except ValueError:
            pass  # e.g. different drive on Windows - just skip showing the image

    # Preserve the current filter/sort query string, so the "Back to History" link
    # and prev/next navigation both keep whatever filters were applied.
    query_string = request.query_string.decode()

    context = base_context()
    context.update({
        "activity": doc,
        "image_url": image_url,
        "prev_id": prev_id,
        "next_id": next_id,
        "query_string": query_string,
    })

    try:
        return render_template("main/activity_detail.html", **context)
    except Exception as e:
        current_app.logger.error(f"Error rendering activity detail page: {e}")
        return "Something went wrong loading activity details.", 500


VALID_ERROR_TYPES = {"SYSTEM_ERROR", "PROCESS_ERROR", "BOTH", "ALL_OK"}


@main_bp.route("/history/activity/<activity_id>/review", methods=["POST"])
def save_activity_review(activity_id):
    """
    Save the review fields for one activity: mark_discuss (YES/NO/untouched -
    a single toggle button in the UI), error_type (SYSTEM_ERROR/PROCESS_ERROR/
    BOTH/ALL_OK/untouched - one consolidated field, replacing the old separate
    mark_ocr_wrong/mark_process_error which could disagree with each other -
    e.g. both marked YES had no defined meaning), and review_comment.

    error_type_marked_by is always set to "VERIFICATION_TEAM" here, since this
    route is only ever reached via a human using this form. The (not yet
    built) AI verification system will set error_type through a different
    path later, with marked_by="AI". Every CHANGE to error_type is appended
    to error_type_history (full audit trail) - re-saving the same value
    doesn't spam a duplicate entry.

    result_reviewed = True if ANY of mark_discuss / error_type / comment has
    been touched (including an explicit "No" or a comment that's since been
    cleared back to blank counts as untouched again).
    """
    db = current_app.extensions.get("mongo_db")
    if db is None:
        return "Database not configured.", 503

    try:
        obj_id = ObjectId(activity_id)
    except InvalidId:
        return "Activity not found.", 404

    mark_discuss = request.form.get("mark_discuss") or None  # "YES" / "NO" / None (untouched)
    error_type = request.form.get("error_type") or None
    if error_type is not None and error_type not in VALID_ERROR_TYPES:
        error_type = None  # ignore anything unexpected rather than store garbage
    review_comment = request.form.get("review_comment", "").strip()

    result_reviewed = bool(mark_discuss or error_type or review_comment)

    update_fields = {
        "mark_discuss": mark_discuss,
        "error_type": error_type,
        "error_type_marked_by": ("VERIFICATION_TEAM" if error_type else None),
        "review_comment": review_comment,
        "result_reviewed": result_reviewed,
    }

    try:
        existing = db["all_activities"].find_one({"_id": obj_id}, {"error_type": 1})
        previous_error_type = existing.get("error_type") if existing else None

        update_op = {"$set": update_fields}

        # Only append a history entry when error_type actually CHANGED -
        # re-submitting the same value (e.g. just editing the comment)
        # shouldn't spam a duplicate audit entry.
        if error_type is not None and error_type != previous_error_type:
            update_op["$push"] = {
                "error_type_history": {
                    "value": error_type,
                    "marked_by": "VERIFICATION_TEAM",
                    "marked_at": datetime.now(timezone.utc).isoformat(),
                    "source": "review",
                }
            }

        db["all_activities"].update_one({"_id": obj_id}, update_op)
        flash("Review saved.", "success")
    except Exception as e:
        current_app.logger.error(f"Failed to save review for activity '{activity_id}': {e}")
        flash("Failed to save review - please try again.", "error")

    # Redirect back to the same detail page, preserving whatever filter/sort
    # query string was active (passed through as a hidden form field).
    query_string = request.form.get("query_string", "")
    target = url_for("main.activity_detail", activity_id=activity_id)
    if query_string:
        target += f"?{query_string}"
    return redirect(target)
